import openpyxl, os

wb = openpyxl.Workbook()

# print(wb.sheetnames)
print(wb['Sheet'])
sheet = wb['Sheet']
print(sheet['A1'].value)

sheet['A1'] = 42
sheet['A2'] = 'Hello'
# Upto this point the data only exists in the computer memory

wb.save('example2.xlsx') #Saves the file in the current folder

sheet2 = wb.create_sheet()
print(wb.sheetnames)
sheet2.title = 'My New Sheet Name'
print(wb.sheetnames)
wb.save('example2.xlsx')
wb.create_sheet(index=0, title='My other sheet')
wb.save('example3.xlsx')