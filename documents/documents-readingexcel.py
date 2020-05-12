import openpyxl, os

workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook))

sheet = workbook['Sheet1']
print(type(sheet))

print(workbook.sheetnames)

print(sheet['A1'].value)
print(sheet['B1'].value)

print((sheet.cell(row=1, column=2)).value)